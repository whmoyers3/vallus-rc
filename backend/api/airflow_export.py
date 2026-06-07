"""Generate the airflow balancing spreadsheet (.xlsx) from a calculated draft.

The sheet mirrors the field balancing form: one tab per HVAC unit, load-share
targets that float against the measured system total, and a hidden reference
sheet holding all 8 orientations x {Cool, Heat, Avg} CFM per room so the two
dropdowns (Orientation, CFM Basis) drive the Load column live via INDEX/MATCH.

See CONTEXT.md -> "Airflow Balancing Export" and ADR 0002.
"""

from __future__ import annotations

import copy
import io
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

from backend.api.serialization import loads_response, project_from_payload
from backend.engine import calculate_project

ORIENTATIONS = ["N", "NE", "E", "SE", "S", "SW", "W", "NW"]
BASES = ["Cool", "Heat", "Avg"]
TONNAGES = ["1.5", "2.0", "2.5", "3.0", "3.5", "4.0", "5.0"]
SYSTEM_TYPES = ["Heat Pump", "AC", "Gas", "Dual Fuel"]
DEFAULT_SYSTEM_TYPE = "Heat Pump"
DEFAULT_BASIS = "Avg"
DEFAULT_ORIENTATION = "S"

GREEN = "578625"
BLUE_INPUT = "489BC9"
GRAY = "D9D9D9"          # calculated-total cells
INPUT_SHADE = "E2EFDA"  # pale-green entry/dropdown cells
ZONE_FILLS = ["DBEAFE", "FEF9C3", "DCFCE7", "FCE7F3", "E0E7FF", "FFEDD5"]
ROOM_ROWS = 22  # supply rows 7..28, matching the single-page form factor
FONT = "Arial"

_green_fill = PatternFill("solid", fgColor=GREEN)
_blue_fill = PatternFill("solid", fgColor=BLUE_INPUT)
_gray_fill = PatternFill("solid", fgColor=GRAY)
_input_fill = PatternFill("solid", fgColor=INPUT_SHADE)
_header_font = Font(name=FONT, bold=True, color="FFFFFF")
_load_font = Font(name=FONT, bold=True, color="FFFFFF")
_label_font = Font(name=FONT, bold=True)
_base_font = Font(name=FONT)
_center = Alignment(horizontal="center", vertical="center")
_thin = Side(style="thin", color="BFBFBF")
_thin_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _grid(ws: Worksheet, c1: int, r1: int, c2: int, r2: int) -> None:
    """Thin border on every cell in the rectangle (merged cells render outer box only)."""
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            ws.cell(row=row, column=col).border = _thin_border


# ── Engine: 8-orientation reference table ─────────────────────────────────────

def _rotate_direction(direction: Any, steps: int) -> Any:
    """Rotate a compass direction clockwise by `steps` 45-degree increments."""
    if direction in ORIENTATIONS:
        return ORIENTATIONS[(ORIENTATIONS.index(direction) + steps) % len(ORIENTATIONS)]
    return direction


def _orientation_table(payload: dict[str, Any]) -> dict[str, dict[str, dict[str, int]]]:
    """room name -> orientation -> {Cool, Heat, Avg} CFM, across all 8 orientations.

    The engine does not read front_door_faces; component directions are already
    resolved to absolute compass headings for the project's *current* orientation.
    To model rotating the house we rotate every surface direction by the angular
    delta between the target orientation and the current one, then recalculate.
    """
    current = (payload["project"].get("metadata") or {}).get("front_door_faces") or DEFAULT_ORIENTATION
    base_idx = ORIENTATIONS.index(current) if current in ORIENTATIONS else ORIENTATIONS.index(DEFAULT_ORIENTATION)
    table: dict[str, dict[str, dict[str, int]]] = {}
    for orient in ORIENTATIONS:
        steps = (ORIENTATIONS.index(orient) - base_idx) % len(ORIENTATIONS)
        variant = copy.deepcopy(payload)
        for level in variant["project"]["levels"]:
            for item in level.get("line_items", []):
                if item.get("direction"):
                    item["direction"] = _rotate_direction(item["direction"], steps)
        result = calculate_project(project_from_payload(variant))
        loads = loads_response(result)
        for level in loads["levels"]:
            for room in level["rooms"]:
                table.setdefault(room["name"], {})[orient] = {
                    "Cool": room["cfm_cool"],
                    "Heat": room["cfm_heat"],
                    "Avg": room["cfm_avg"],
                }
    return table


# ── Unit / zone grouping (from the input payload) ─────────────────────────────

def _group_units(payload: dict[str, Any]) -> list[dict[str, Any]]:
    """Return [{id, name, rooms:[{name, zone_id}], zones:[{id,name}]}] ordered."""
    proj = payload["project"]
    meta = proj.get("metadata") or {}
    units = meta.get("units") or [{"id": "__default__", "name": proj.get("name") or "Whole House"}]
    zone_defs = {z.get("id"): z.get("name") for z in (meta.get("zones") or [])}

    rooms: list[dict[str, Any]] = []
    for level in proj["levels"]:
        for r in level.get("rooms", []):
            rooms.append({"name": r["name"], "unit_id": r.get("unit_id"), "zone_id": r.get("zone_id")})

    any_unit_assigned = any(r["unit_id"] for r in rooms)
    grouped: list[dict[str, Any]] = []
    for idx, unit in enumerate(units):
        uid = unit.get("id")
        if any_unit_assigned:
            unit_rooms = [r for r in rooms if r["unit_id"] == uid]
        else:
            unit_rooms = rooms if idx == 0 else []
        if not unit_rooms:
            continue
        # zone order of appearance within the unit
        zone_order: list[str] = []
        for r in unit_rooms:
            zid = r["zone_id"]
            if zid is not None and zid not in zone_order:
                zone_order.append(zid)
        grouped.append({
            "id": uid,
            "name": unit.get("name") or "Unit",
            "rooms": unit_rooms,
            "zone_order": zone_order,
            "zone_names": zone_defs,
            "selected_tons": float(proj["levels"][0].get("selected_tons") or 0),
        })
    return grouped


def _zone_fill_for(zone_id: Any, zone_order: list[str]) -> PatternFill | None:
    if zone_id is None or zone_id not in zone_order:
        return None
    return PatternFill("solid", fgColor=ZONE_FILLS[zone_order.index(zone_id) % len(ZONE_FILLS)])


# ── Hidden reference sheet ────────────────────────────────────────────────────

def _build_ref_sheet(wb: Workbook, table: dict[str, dict[str, dict[str, int]]]) -> int:
    ws = wb.create_sheet("Ref")
    ws.sheet_state = "hidden"
    ws["A1"] = "Area"
    col = 2
    for orient in ORIENTATIONS:
        for basis in BASES:
            ws.cell(row=1, column=col, value=f"{orient}_{basis}")
            col += 1
    row = 2
    for room_name, by_orient in table.items():
        ws.cell(row=row, column=1, value=room_name)
        col = 2
        for orient in ORIENTATIONS:
            for basis in BASES:
                ws.cell(row=row, column=col, value=by_orient.get(orient, {}).get(basis, 0))
                col += 1
        row += 1
    return row - 1  # last data row


# ── Unit tab ──────────────────────────────────────────────────────────────────

def _add_validation(ws: Worksheet, cell: str, options: list[str]) -> None:
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', showDropDown=False, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws[cell])


def _build_unit_sheet(
    wb: Workbook,
    unit: dict[str, Any],
    ref_last_row: int,
    address: str,
    plan_label: str,
    default_orientation: str,
) -> None:
    title = re.sub(r"[\/?*\[\]:]", "-", unit["name"])[:31] or "Unit"
    ws = wb.create_sheet(title)
    ws.sheet_view.zoomScale = 110
    ws.sheet_view.showGridLines = False
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    zone_names = unit["zone_names"]
    zone_label = ""
    if unit["zone_order"]:
        labels = [f"Zone {chr(65 + i)}: {zone_names.get(z, z)}" for i, z in enumerate(unit["zone_order"])]
        zone_label = " (" + ", ".join(labels) + ")"

    # ── Header block (rows 1-4): left labels A + values B:G; right labels I:J + values K:L
    for r in range(1, 5):
        ws.merge_cells(f"B{r}:G{r}")
        ws.merge_cells(f"I{r}:J{r}")
        ws.merge_cells(f"K{r}:L{r}")
    ws["A1"], ws["A2"], ws["A3"], ws["A4"] = "Where", "Plan", "Unit", "Orientation"
    ws["B1"] = address
    ws["B2"] = plan_label
    ws["B3"] = f"{unit['name']}{zone_label}"
    ws["B4"] = default_orientation
    ws["I1"], ws["K1"] = "Test Taken", "=TODAY()"
    ws["I2"], ws["K2"] = "System Type", DEFAULT_SYSTEM_TYPE
    ws["I3"], ws["K3"] = "Capacity", unit["selected_tons"]
    ws["I4"], ws["K4"] = "CFM Basis", DEFAULT_BASIS
    ws["K3"].number_format = "0.0"
    for cell in ("A1", "A2", "A3", "A4", "I1", "I2", "I3", "I4"):
        ws[cell].font = _label_font
    for cell in ("B1", "B2", "B3", "K1"):
        ws[cell].font = _base_font
    for cell in ("B4", "K2", "K3", "K4"):  # entry/dropdown cells
        ws[cell].fill = _input_fill
    _grid(ws, 1, 1, 7, 4)    # left block A..G
    _grid(ws, 9, 1, 12, 4)   # right block I..L
    _add_validation(ws, "K2", SYSTEM_TYPES)
    _add_validation(ws, "K3", TONNAGES)
    _add_validation(ws, "B4", ORIENTATIONS)
    _add_validation(ws, "K4", BASES)

    # ── Supply table header (row 6): A | B:E readings | F Total | G Load | H % | I Adj Target | J CFM+/- | K %+/-
    ws.merge_cells("B6:E6")
    supply_headers = {
        "A6": "Room", "B6": "Supply Actual Readings", "F6": "Total", "G6": "Load",
        "H6": "%", "I6": "Adj  Target", "J6": "CFM +/-", "K6": "% +/-",
    }
    for cell, val in supply_headers.items():
        ws[cell] = val
        ws[cell].fill = _green_fill
        ws[cell].font = _header_font
        ws[cell].alignment = wrap_center
    ws.row_dimensions[6].height = 29

    data_start = 7
    n_rows = max(ROOM_ROWS, len(unit["rooms"]))
    data_end = data_start + n_rows - 1
    total_row = data_end + 1

    ref_key_row = "Ref!$B$1:$Y$1"
    ref_name_col = f"Ref!$A$2:$A${ref_last_row}"
    ref_val_range = f"Ref!$B$2:$Y${ref_last_row}"

    for offset in range(n_rows):
        r = data_start + offset
        ws[f"F{r}"] = f'=IF(SUM(B{r}:E{r})=0,"",SUM(B{r}:E{r}))'
        ws[f"F{r}"].fill = _gray_fill
        ws[f"H{r}"] = f'=IF(OR(G{r}="",$G${total_row}=0),"",G{r}/$G${total_row})'
        ws[f"I{r}"] = f'=IF(OR(F{r}="",H{r}=""),"",$F${total_row}*H{r})'
        ws[f"J{r}"] = f'=IF(OR(G{r}="",F{r}="",I{r}=0),"",F{r}-I{r})'
        # % the room must move to reach target, relative to its current reading
        ws[f"K{r}"] = f'=IF(OR(I{r}="",F{r}="",F{r}=0),"",(F{r}-I{r})/F{r})'
        ws[f"H{r}"].number_format = "0%"   # whole percent
        ws[f"I{r}"].number_format = "0"    # whole CFM
        ws[f"J{r}"].number_format = "0"    # whole CFM (signed)
        ws[f"K{r}"].number_format = "0%"   # whole percent

        zone_fill = None
        if offset < len(unit["rooms"]):
            room = unit["rooms"][offset]
            ws[f"A{r}"] = room["name"]
            ws[f"G{r}"] = (
                f'=IFERROR(INDEX({ref_val_range},'
                f'MATCH($A{r},{ref_name_col},0),'
                f'MATCH($B$4&"_"&$K$4,{ref_key_row},0)),"")'
            )
            ws[f"G{r}"].fill = _blue_fill
            ws[f"G{r}"].font = _load_font
            ws[f"G{r}"].alignment = _center
            zone_fill = _zone_fill_for(room["zone_id"], unit["zone_order"])
        if zone_fill is not None:
            for col in range(1, 12):  # A..K
                if col in (6, 7):     # skip F (gray total) and G (blue load)
                    continue
                ws.cell(row=r, column=col).fill = zone_fill

    # System totals
    ws[f"F{total_row}"] = f"=SUM(F{data_start}:F{data_end})"
    ws[f"G{total_row}"] = f"=SUM(G{data_start}:G{data_end})"
    ws[f"F{total_row}"].fill = _gray_fill
    ws[f"F{total_row}"].font = _label_font
    ws[f"G{total_row}"].font = _label_font
    ws[f"G{total_row}"].alignment = _center
    _grid(ws, 1, 6, 11, total_row)  # box supply table A6:K{total}

    # Conditional format: red when % +/- outside +-10%
    red_fill = PatternFill("solid", fgColor="FECACA")
    ws.conditional_formatting.add(
        f"K{data_start}:K{data_end}",
        FormulaRule(formula=[f'AND(K{data_start}<>"",OR(K{data_start}>0.1,K{data_start}<-0.1))'], fill=red_fill),
    )

    # ── Lower blocks (rh = total_row + 2) ──
    rh = total_row + 2
    grand = rh + 9
    # Return-air table: A | B:D readings | E total
    ws.merge_cells(f"B{rh}:D{rh}")
    ws.merge_cells(f"J{rh}:K{rh}")
    ws[f"A{rh}"], ws[f"B{rh}"], ws[f"E{rh}"] = "Room", "Return Actuals", "Total"
    ws[f"J{rh}"], ws[f"L{rh}"] = "Check List", "(y/n)"
    for cell in (f"A{rh}", f"B{rh}", f"E{rh}", f"J{rh}", f"K{rh}", f"L{rh}"):
        ws[cell].fill = _green_fill
        ws[cell].font = _header_font
        ws[cell].alignment = _center
    for i in range(8):  # white return reading rows; gray E total
        r = rh + 1 + i
        ws[f"E{r}"] = f'=IF(SUM(B{r}:D{r})=0,"",SUM(B{r}:D{r}))'
        ws[f"E{r}"].fill = _gray_fill
    ws[f"E{grand}"] = f'=IF(SUM(E{rh + 1}:E{rh + 8})=0,"",SUM(E{rh + 1}:E{rh + 8}))'
    ws[f"E{grand}"].fill = _gray_fill
    ws[f"E{grand}"].font = _label_font
    _grid(ws, 1, rh, 5, grand)

    # Static pressure (G:H merged), gap row at rh+6 between inputs and outputs
    gap = rh + 6
    for r in range(rh, rh + 11):
        if r != gap:
            ws.merge_cells(f"G{r}:H{r}")
    ws[f"G{rh}"] = "Supply"
    ws[f"G{rh + 2}"] = "Return"
    ws[f"G{rh + 4}"] = "Before Filter"
    ws[f"G{rh + 7}"] = "Total"
    ws[f"G{rh + 8}"] = f'=IF(G{rh + 1}=0,"",ABS(G{rh + 3})+ABS(G{rh + 1}))'
    ws[f"G{rh + 9}"] = "Filter Drop"
    ws[f"G{rh + 10}"] = f'=IF(G{rh + 5}=0,"",ABS(G{rh + 3})-ABS(G{rh + 5}))'
    for cell in (f"G{rh}", f"G{rh + 2}", f"G{rh + 4}", f"G{rh + 7}", f"G{rh + 9}"):
        ws[cell].font = _label_font
    for cell in (f"G{rh + 1}", f"G{rh + 3}", f"G{rh + 5}"):  # measured input cells
        ws[cell].fill = _input_fill
    _grid(ws, 7, rh, 8, gap - 1)
    _grid(ws, 7, gap + 1, 8, rh + 10)

    # Filter Type label + input
    ws[f"K{rh + 8}"] = "Filter Type"
    ws[f"K{rh + 8}"].font = _label_font
    ws.merge_cells(f"K{rh + 9}:L{rh + 9}")
    ws[f"K{rh + 9}"].fill = _input_fill
    _grid(ws, 11, rh + 8, 12, rh + 9)

    # Go/No-Go checklist: labels merged J:K, Y/N input in L
    checklist = ["Size Match", "Total Airflow", "Room-Room", "Strip check", "Cool Check", "Zone Check"]
    for i, label in enumerate(checklist):
        r = rh + 1 + i
        ws.merge_cells(f"J{r}:K{r}")
        ws[f"J{r}"] = label
        ws[f"J{r}"].alignment = _center           # merged + centered labels
        ws[f"L{r}"].fill = _input_fill
        ws[f"L{r}"].alignment = _center            # centered y/n input
    _grid(ws, 10, rh, 12, rh + len(checklist))  # J..L

    # Notes box
    notes_label = grand + 2
    ws[f"A{notes_label}"] = "Notes"
    ws[f"A{notes_label}"].font = _label_font
    ws.merge_cells(f"B{notes_label + 1}:K{notes_label + 4}")
    _grid(ws, 2, notes_label + 1, 11, notes_label + 4)

    # ── Spacing: column widths + row heights (preserve single-page layout) ──
    # Reading columns B-E share the same total span (~43.3) but split evenly so
    # every cell is wide enough to type into (column B was awkwardly thin).
    widths = {"A": 14.66, "B": 10.83, "C": 10.83, "D": 10.83, "E": 10.83, "F": 5.33,
              "G": 6.0, "H": 6.33, "I": 7.5, "J": 6.33, "K": 6.5, "L": 7.16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for r in range(data_start, data_end + 1):
        ws.row_dimensions[r].height = 16.5
    ws.freeze_panes = "B7"

    # ── Print: fit to one page WIDE, natural height. fitToHeight is left at 0
    # (unlimited) so cell sizing stays identical across exports regardless of
    # room count; only the column-fit scaling (constant, same columns always)
    # applies. Typical homes still land on a single page.
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5


def _build_zone_key(wb: Workbook, units: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Zone Key")
    ws["A1"] = "Zone Color Key"
    ws["A1"].font = _label_font
    row = 3
    for unit in units:
        if not unit["zone_order"]:
            continue
        ws.cell(row=row, column=1, value=unit["name"]).font = _label_font
        row += 1
        for i, zid in enumerate(unit["zone_order"]):
            cell = ws.cell(row=row, column=1, value=f"Zone {chr(65 + i)}: {unit['zone_names'].get(zid, zid)}")
            cell.fill = PatternFill("solid", fgColor=ZONE_FILLS[i % len(ZONE_FILLS)])
            row += 1
        row += 1
    if row == 3:
        ws["A3"] = "No zones defined for this project."
    ws.column_dimensions["A"].width = 40


# ── Public entry point ────────────────────────────────────────────────────────

def _plan_label(payload: dict[str, Any]) -> str:
    meta = payload["project"].get("metadata") or {}
    src = meta.get("source_filename")
    if src:
        return re.sub(r"\.pdf$", "", src, flags=re.IGNORECASE)
    return payload["project"].get("description") or payload["project"].get("name") or "Plan"


def _safe_filename(name: str) -> str:
    base = re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("_") or "Project"
    return f"{base}_Airflow.xlsx"


def build_airflow_workbook(payload: dict[str, Any]) -> tuple[bytes, str]:
    """Build the balancing workbook. Returns (xlsx_bytes, filename)."""
    meta = payload["project"].get("metadata") or {}
    default_orientation = meta.get("front_door_faces") or DEFAULT_ORIENTATION
    address = meta.get("address") or ""
    plan_label = _plan_label(payload)

    table = _orientation_table(payload)
    units = _group_units(payload)

    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet
    ref_last_row = _build_ref_sheet(wb, table)
    for unit in units:
        _build_unit_sheet(wb, unit, ref_last_row, address, plan_label, default_orientation)
    _build_zone_key(wb, units)
    # Ref hidden, so make first unit tab active
    wb.active = 1

    buffer = io.BytesIO()
    wb.save(buffer)
    plan_name = re.sub(r"\.pdf$", "", (meta.get("source_filename") or payload["project"].get("name") or "Project"), flags=re.IGNORECASE)
    return buffer.getvalue(), _safe_filename(plan_name)
